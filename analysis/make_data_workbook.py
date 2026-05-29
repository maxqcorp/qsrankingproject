"""Build a single Excel workbook with the QS data on four sheets:
2023, 2024, 2025, and a combined long-format sheet across the three editions.

Source data are the official QS workbooks parsed by qs_common.load_year.
Output: data/QS_data_2023_2025.xlsx
"""
import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from qs_common import load_year, FILES, IND, IND_LABEL, ROOT

RENAME = {"institution": "Institution", "country": "Country", "region": "Region",
          "research": "Research intensity", "size": "Size", "rank_num": "Rank",
          "overall_pub": "Overall score (published)"}
RENAME.update({k: IND_LABEL[k] for k in IND})
COL_ORDER = ["institution", "country", "region", "research", "size", "rank_num"] + IND + ["overall_pub"]


def tidy(year):
    return load_year(year)[COL_ORDER].rename(columns=RENAME)


def main():
    frames = {y: tidy(y) for y in FILES}
    combined = pd.concat([f.assign(Year=y) for y, f in frames.items()], ignore_index=True)
    combined = combined[["Year"] + [c for c in combined.columns if c != "Year"]]

    data_dir = os.path.join(ROOT, "data")
    os.makedirs(data_dir, exist_ok=True)
    path = os.path.join(data_dir, "QS_data_2023_2025.xlsx")

    sheets = {str(y): frames[y] for y in FILES}
    sheets["Combined 2023-2025"] = combined
    with pd.ExcelWriter(path, engine="openpyxl") as xw:
        for name, df in sheets.items():
            df.to_excel(xw, sheet_name=name, index=False)

    wb = load_workbook(path)
    head_fill = PatternFill("solid", fgColor="264653")
    head_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    body_font = Font(name="Arial", size=10)
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.fill = head_fill
            cell.font = head_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.font = body_font
        for col_idx in range(1, ws.max_column + 1):
            header = str(ws.cell(row=1, column=col_idx).value or "")
            width = 42 if header == "Institution" else max(9, min(20, len(header) + 2))
            ws.column_dimensions[get_column_letter(col_idx)].width = width
        ws.row_dimensions[1].height = 30
    wb.save(path)

    print(f"wrote {path}")
    for name, df in sheets.items():
        print(f"  sheet '{name}': {df.shape[0]} rows x {df.shape[1]} cols")


if __name__ == "__main__":
    main()
